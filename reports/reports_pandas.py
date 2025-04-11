# Exporta o relatorio de despesas para XLSX formatado com Excel

import pandas as pd # Importa a biblioteca pandas para manipulação de dados
from sqlalchemy import create_engine
from urllib.parse import quote_plus  # Importa para codificar a senha corretamente
from database.db_config import get_db_connection # Importa a função para obter a conexão com o banco de dados
from datetime import datetime # Importa a classe datetime para manipulação de data e hora
import traceback # Para exibir erros detalhados
import sys # Para encerrar o programa corretamente
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows

def exportar_para_excel(usuario_id, mes_ano):
    """
    Exporta as despesas de um usuário específico
    para CSV, filtrando por mês.
    """
    try: # Alguma operação que pode gerar erro
        
        def get_db_connection():
            """
            Cria e retorna uma conexão 
            segura com o banco de dados.
            """
            password = quote_plus("@program225X")  # Codifica a senha para evitar erros com caracteres especiais
            engine = create_engine(f"mysql+pymysql://Vinicius:{password}@localhost/gestao_despesas",
                           pool_pre_ping=True,  # Evita erros de conexão perdida
                           connect_args={"charset": "utf8mb4"})  # Garante suporte a caracteres especiais
            return engine.connect()

        query = """
                SELECT c.nome AS Categoria, d.valor, d.data, d.descricao, d.fixa,
                        (SELECT hr.renda_mensal FROM historico_renda hr
                        WHERE hr.usuario_id = d.usuario_id
                        AND DATE_FORMAT(hr.data_registro, '%%Y-%%m') = %s
                        ORDER BY hr.data_registro DESC LIMIT 1) AS renda_mensal
                FROM despesas d  
                JOIN categorias c ON d.categoria_id = c.categoria_id
                WHERE d.usuario_id = %s AND DATE_FORMAT(d.data, '%%Y-%%m') = %s
            """
            
        # Criando conexão e executando a query
        with get_db_connection() as conn: # Isso garante que a conexão seja fechada automaticamente
            df = pd.read_sql_query(query, conn, params=(mes_ano, usuario_id, mes_ano)) # Executa a consulta e armazena o resultado em um DataFrame
            
        if df.empty:
            print("Nenhuma despesa encontrada para exportar.")
            return
              
        print("Colunas disponíveis no DataFrame: ", df.columns)
            
        # Converte 1 para "Fixa" e 0 para "Variável"
        if 'fixa' in df.columns:
            df['fixa'] = df['fixa'].map({1: 'Fixa', 0: 'Variável'}) # Converte antes de renomear
        else:
                print("Atenção: A coluna 'fixa' não foi encontrada no banco de dados. O tipo de despesa não será incluído.")
            
        # Verifica se há dados antes de gerar o gráfico
        df.rename(columns={
            'renda_mensal': 'Renda Mensal',
            'valor': 'Valor',
            'data': 'Data',
            'descricao': 'Descrição',
            'fixa': 'Tipo'
        }, inplace=True)
        
        df['Data'] = pd.to_datetime(df['Data'])
                        
        # Ordena por data
        df.sort_values(by='Data', inplace=True)
        
        # Extrai a renda mensal (assumindo que é igual em todas as linhas)
        renda_mensal_valor = df['Renda Mensal'].iloc[0] if 'Renda Mensal' in df.columns else None
        
        # Remove a coluna "Renda Mensal" do DataFrame
        if 'Renda Mensal' in df.columns:
            df.drop(columns=['Renda Mensal'], inplace=True)
        
        data_atual = datetime.now().strftime("%Y-%m-%d")
        nome_arquivo = f"relatorio_despesas_usuario_{usuario_id}_{mes_ano}_{data_atual}.xlsx" 
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório de Despesas"
        
        # Formatação
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center")
        money_format = '#,##0.00\ [$R$-416]'
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        fill = PatternFill("solid", fgColor="DDDDDD")
        
        # Adiciona renda mensal
        if renda_mensal_valor is not None:
            ws.append(["Renda Mensal"]) # Título na primeira linha
            ws.append([renda_mensal_valor]) # Valor na linha de baixo
            
            ws["A1"].font = bold_font
            ws["A1"].alignment = center_align
            ws["A1"].border = border
            ws["A1"].fill = fill
            
            renda_cell = ws["A2"]
            renda_cell.number_format = money_format
            renda_cell.font = Font(bold=True)
            renda_cell.alignment = center_align
            renda_cell.border = border
            
            ws.append([]) # Linha em branco

        # Adiciona os dados do DataFrame
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
            
        # Determina em que linha começam os dados
        linha_inicio_dados = 4 if renda_mensal_valor is not None else 2
        
        # Cabeçalho formatado
        for col in ws.iter_cols(min_row=linha_inicio_dados, max_row=linha_inicio_dados, min_col=1, max_col=ws.max_column):
            for cell in col:
                cell.font = bold_font
                cell.fill = fill
                cell.border = border
                cell.alignment = center_align
          
        # Aplica formato de moeda na coluna "Valor"
        for row in ws.iter_rows(min_row=linha_inicio_dados + 1, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                # Formata valor
                cell.border = border
                if cell.column_letter == 'B' and isinstance(cell.value, (int, float)):
                    cell.number_format = money_format
                    cell.alignment = center_align
                    
                # Formata data (coluna C)
                elif cell.column_letter == 'C' and isinstance(cell.value, datetime):
                    cell.number_format = 'DD/MM/YYYY'
                    cell.alignment = center_align
                
                elif cell.column_letter in ['A', 'D', 'E']:
                    cell.alignment = center_align

        # Ajusta largura das colunas automaticamente
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2
        
        # Força largura maior para a coluna "Valor" (coluna B)
        ws.column_dimensions["B"].width = 15
        
        wb.save(nome_arquivo)   
        print(f"Relatório exportado com sucesso para {nome_arquivo}")
             
    except Exception:
        print("Erro ao exportar relatório para CSV:")
        traceback.print_exc() # Exibe o rastreamento completo do erro
        
if __name__ == "__main__": # Verifica se o script está sendo executado diretamente
    try:
        usuario_id = int(input("Digite o ID do usuário:"))
        mes_ano = input("Digite o mês no formato YYYY-MM (exemplo: 2025-03): ")
        exportar_para_excel(usuario_id, mes_ano)
    except ValueError:
        print("ID do usuário inválido. Certifique-se de inserir um número inteiro.")
        sys.exit(1)